import frappe
from frappe.utils import get_first_day, get_last_day, today

@frappe.whitelist()
def sync_now():
    from vera_drive.tasks import sync_drive_files
    sync_drive_files()
    return {'status': 'ok'}

@frappe.whitelist()
def get_dashboard_stats():
    ms = get_first_day(today())
    me = get_last_day(today())
    filters = {'file_date': ['between', [ms, me]]}
    return {
        'total': frappe.db.count('VE Drive File', filters),
        'pending': frappe.db.count('VE Drive File',
            {**filters, 'status': 'New'}),
        'flagged': frappe.db.count('VE Drive File',
            {**filters, 'status': 'Flagged'}),
        'last_sync': str(frappe.db.get_value(
            'VE Drive Sync Log',
            {'sync_status': 'Success'},
            'sync_time',
            order_by='sync_time desc'
        ) or 'Never')
    }

@frappe.whitelist()
def get_all_files(category='All', status=None):
    filters = {}
    if category and category != 'All':
        filters['category'] = category
    if status:
        filters['status'] = status
    return frappe.get_all('VE Drive File',
        filters=filters,
        fields=[
            'name', 'file_name', 'doc_type', 'category',
            'party_name', 'file_date', 'status',
            'drive_view_link', 'drive_file_id',
            'drive_folder_path', 'file_extension',
            'synced_on', 'admin_notes',
            'uploaded_by_name', 'uploaded_by_email',
            'last_modified_by_name', 'last_modified_by_email',
            'upload_detected_method'
        ],
        order_by='file_date desc, synced_on desc'
    )


@frappe.whitelist()
def get_folder_tree():
    from vera_drive.google_drive import (
        get_drive_service,
        list_files_in_folder,
        map_to_vera_employee,
        ROOT_FOLDER_ID
    )

    service = get_drive_service()

    def _uploader_from_item(item):
        owners = item.get('owners', [])
        last_modifier = item.get('lastModifyingUser', {})
        if owners:
            raw_email = owners[0].get('emailAddress', '')
            vera_email, vera_name = map_to_vera_employee(raw_email)
            if vera_email:
                return vera_name, vera_email, 'drive_api'
            return owners[0].get('displayName', ''), raw_email, 'drive_api_unknown'
        if last_modifier:
            raw_email = last_modifier.get('emailAddress', '')
            vera_email, vera_name = map_to_vera_employee(raw_email)
            if vera_email:
                return vera_name, vera_email, 'last_modifier'
        return '', '', ''

    def build_tree(folder_id, folder_name, path):
        items = list_files_in_folder(service, folder_id)
        folders = []
        files = []
        for item in items:
            if item['mimeType'] == 'application/vnd.google-apps.folder':
                subtree = build_tree(
                    item['id'],
                    item['name'],
                    path + '/' + item['name']
                )
                folders.append(subtree)
            else:
                name = item['name']
                last_modifier = item.get('lastModifyingUser', {})
                uploader_name, uploader_email, detected_method = _uploader_from_item(item)
                files.append({
                    'id': item['id'],
                    'name': name,
                    'mimeType': item['mimeType'],
                    'webViewLink': item.get('webViewLink', ''),
                    'modifiedTime': item.get('modifiedTime', ''),
                    'size': item.get('size', 0),
                    'extension': name.rsplit('.', 1)[-1].lower() if '.' in name else '',
                    'uploaded_by_name': uploader_name,
                    'uploaded_by_email': uploader_email,
                    'upload_detected_method': detected_method,
                    'last_modified_by_name': last_modifier.get('displayName', ''),
                    'last_modified_by_email': last_modifier.get('emailAddress', ''),
                })
        return {
            'id': folder_id,
            'name': folder_name,
            'path': path,
            'type': 'folder',
            'folders': sorted(folders, key=lambda x: x['name']),
            'files': sorted(files, key=lambda x: x['name']),
            'file_count': len(files),
            'total_count': len(files) + sum(f['total_count'] for f in folders)
        }

    return build_tree(
        ROOT_FOLDER_ID,
        'Vera Enterprises \u2014 Documents',
        'Vera Enterprises \u2014 Documents'
    )


@frappe.whitelist()
def get_folder_contents(folder_id):
    from vera_drive.google_drive import (
        get_drive_service,
        list_files_in_folder
    )

    service = get_drive_service()
    items = list_files_in_folder(service, folder_id)
    folders = []
    files = []
    for item in items:
        if item['mimeType'] == 'application/vnd.google-apps.folder':
            folders.append({
                'id': item['id'],
                'name': item['name'],
                'type': 'folder',
            })
        else:
            name = item['name']
            files.append({
                'id': item['id'],
                'name': name,
                'mimeType': item['mimeType'],
                'webViewLink': item.get('webViewLink', ''),
                'modifiedTime': item.get('modifiedTime', ''),
                'size': item.get('size', 0),
                'extension': name.rsplit('.', 1)[-1].lower() if '.' in name else ''
            })
    return {
        'folders': sorted(folders, key=lambda x: x['name']),
        'files': sorted(files, key=lambda x: x['name'])
    }

@frappe.whitelist()
def mark_reviewed(docname):
    frappe.db.set_value(
        'VE Drive File', docname, 'status', 'Reviewed'
    )
    frappe.db.commit()
    return {'status': 'ok'}

@frappe.whitelist()
def flag_file(docname, notes=''):
    frappe.db.set_value(
        'VE Drive File',
        docname,
        {'status': 'Flagged', 'admin_notes': notes}
    )
    frappe.db.commit()
    return {'status': 'ok'}

@frappe.whitelist()
def analyse_file(drive_file_id, file_extension):
    from vera_drive.google_drive import download_file_content
    fh = download_file_content(drive_file_id)
    if file_extension in ['xlsx', 'xls']:
        import openpyxl
        wb = openpyxl.load_workbook(fh, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(max_row=50, values_only=True):
            if any(c is not None for c in row):
                rows.append([
                    str(c) if c is not None else ''
                    for c in row
                ])
        return {
            'type': 'spreadsheet',
            'rows': rows[:30],
            'total_rows': ws.max_row
        }
    elif file_extension == 'pdf':
        import pdfplumber
        with pdfplumber.open(fh) as pdf:
            text = ''.join(
                p.extract_text() or ''
                for p in pdf.pages[:3]
            )
        return {'type': 'pdf', 'text': text[:3000]}
    return {'type': 'unknown'}
